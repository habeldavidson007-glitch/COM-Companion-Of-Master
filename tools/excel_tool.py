"""
COM SGMA LIGHT - Excel Tool
===========================
Phase 2a: Actual pandas-based Excel generation with design system integration.
Integrated color systems: Tailwind CSS, Material Design, Ant Design, IBM Carbon, 
GitHub Primer, Radix UI, Chakra UI, Atlassian, and Open Color.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils.dataframe import dataframe_to_rows
import os


# ============================================================================
# COMPREHENSIVE DESIGN TOKENS - Integrated from Multiple Design Systems
# Matches PDF and PPT tools for cross-format consistency
# ============================================================================

# Brand Core - COM SGMA LIGHT
BRAND_COLORS = {
    "primary": "#1a5f3c",      # Emerald green - primary actions, headers
    "secondary": "#0d3d26",    # Dark emerald - accents, borders
    "accent": "#ffb1ee",       # Soft pink - highlights, CTAs
    "background": "#ffffff",   # White - page background
    "surface": "#f8f9fa",      # Light gray - content blocks
    "text_primary": "#1a1a1a", # Near black - main text
    "text_secondary": "#666666", # Gray - secondary text
    "border": "#e0e0e0",       # Light border
}

# Tailwind CSS Colors - Industry Standard (120K+ stars)
TAILWIND_COLORS = {
    "red": {"50": "#fef2f2", "100": "#fee2e2", "200": "#fecaca", "300": "#fca5a5", "400": "#f87171", "500": "#ef4444", "600": "#dc2626", "700": "#b91c1c", "800": "#991b1b", "900": "#7f1d1d"},
    "orange": {"50": "#fff7ed", "100": "#ffedd5", "200": "#fed7aa", "300": "#fdba74", "400": "#fb923c", "500": "#f97316", "600": "#ea580c", "700": "#c2410c", "800": "#9a3412", "900": "#7c2d12"},
    "amber": {"50": "#fffbeb", "100": "#fef3c7", "200": "#fde68a", "300": "#fcd34d", "400": "#fbbf24", "500": "#f59e0b", "600": "#d97706", "700": "#b45309", "800": "#92400e", "900": "#78350f"},
    "yellow": {"50": "#fefce8", "100": "#fef9c3", "200": "#fef08a", "300": "#fde047", "400": "#facc15", "500": "#eab308", "600": "#ca8a04", "700": "#a16207", "800": "#854d0e", "900": "#713f12"},
    "green": {"50": "#f0fdf4", "100": "#dcfce7", "200": "#bbf7d0", "300": "#86efac", "400": "#4ade80", "500": "#22c55e", "600": "#16a34a", "700": "#15803d", "800": "#166534", "900": "#14532d"},
    "emerald": {"50": "#ecfdf5", "100": "#d1fae5", "200": "#a7f3d0", "300": "#6ee7b7", "400": "#34d399", "500": "#10b981", "600": "#059669", "700": "#047857", "800": "#065f46", "900": "#064e3b"},
    "teal": {"50": "#f0fdfa", "100": "#ccfbf1", "200": "#99f6e4", "300": "#5eead4", "400": "#2dd4bf", "500": "#14b8a6", "600": "#0d9488", "700": "#0f766e", "800": "#115e59", "900": "#134e4a"},
    "cyan": {"50": "#ecfeff", "100": "#cffafe", "200": "#a5f3fc", "300": "#67e8f9", "400": "#22d3ee", "500": "#06b6d4", "600": "#0891b2", "700": "#0e7490", "800": "#155e75", "900": "#164e63"},
    "sky": {"50": "#f0f9ff", "100": "#e0f2fe", "200": "#bae6fd", "300": "#7dd3fc", "400": "#38bdf8", "500": "#0ea5e9", "600": "#0284c7", "700": "#0369a1", "800": "#075985", "900": "#0c4a6e"},
    "blue": {"50": "#eff6ff", "100": "#dbeafe", "200": "#bfdbfe", "300": "#93c5fd", "400": "#60a5fa", "500": "#3b82f6", "600": "#2563eb", "700": "#1d4ed8", "800": "#1e40af", "900": "#1e3a8a"},
    "indigo": {"50": "#eef2ff", "100": "#e0e7ff", "200": "#c7d2fe", "300": "#a5b4fc", "400": "#818cf8", "500": "#6366f1", "600": "#4f46e5", "700": "#4338ca", "800": "#3730a3", "900": "#312e81"},
    "violet": {"50": "#f5f3ff", "100": "#ede9fe", "200": "#ddd6fe", "300": "#c4b5fd", "400": "#a78bfa", "500": "#8b5cf6", "600": "#7c3aed", "700": "#6d28d9", "800": "#5b21b6", "900": "#4c1d95"},
    "purple": {"50": "#faf5ff", "100": "#f3e8ff", "200": "#e9d5ff", "300": "#d8b4fe", "400": "#c084fc", "500": "#a855f7", "600": "#9333ea", "700": "#7e22ce", "800": "#6b21a8", "900": "#581c87"},
    "pink": {"50": "#fdf2f8", "100": "#fce7f3", "200": "#fbcfe8", "300": "#f9a8d4", "400": "#f472b6", "500": "#ec4899", "600": "#db2777", "700": "#be185d", "800": "#9d174d", "900": "#831843"},
    "rose": {"50": "#fff1f2", "100": "#ffe4e6", "200": "#fecdd3", "300": "#fda4af", "400": "#fb7185", "500": "#f43f5e", "600": "#e11d48", "700": "#be123c", "800": "#9f1239", "900": "#881337"},
    "slate": {"50": "#f8fafc", "100": "#f1f5f9", "200": "#e2e8f0", "300": "#cbd5e1", "400": "#94a3b8", "500": "#64748b", "600": "#475569", "700": "#334155", "800": "#1e293b", "900": "#0f172a"},
    "gray": {"50": "#f9fafb", "100": "#f3f4f6", "200": "#e5e7eb", "300": "#d1d5db", "400": "#9ca3af", "500": "#6b7280", "600": "#4b5563", "700": "#374151", "800": "#1f2937", "900": "#111827"},
}

# Material Design Colors - Google's Accessibility-Tested System (40K+ stars)
MATERIAL_COLORS = {
    "red": {"50": "#ffebee", "100": "#ffcdd2", "200": "#ef9a9a", "300": "#e57373", "400": "#ef5350", "500": "#f44336", "600": "#e53935", "700": "#d32f2f", "800": "#c62828", "900": "#b71c1c"},
    "pink": {"50": "#fce4ec", "100": "#f8bbd0", "200": "#f48fb1", "300": "#f06292", "400": "#ec407a", "500": "#e91e63", "600": "#d81b60", "700": "#c2185b", "800": "#ad1457", "900": "#880e4f"},
    "purple": {"50": "#f3e5f5", "100": "#e1bee7", "200": "#ce93d8", "300": "#ba68c8", "400": "#ab47bc", "500": "#9c27b0", "600": "#8e24aa", "700": "#7b1fa2", "800": "#6a1b9a", "900": "#4a148c"},
    "blue": {"50": "#e3f2fd", "100": "#bbdefb", "200": "#90caf9", "300": "#64b5f6", "400": "#42a5f5", "500": "#2196f3", "600": "#1e88e5", "700": "#1976d2", "800": "#1565c0", "900": "#0d47a1"},
    "cyan": {"50": "#e0f7fa", "100": "#b2ebf2", "200": "#80deea", "300": "#4dd0e1", "400": "#26c6da", "500": "#00bcd4", "600": "#00acc1", "700": "#0097a7", "800": "#00838f", "900": "#006064"},
    "teal": {"50": "#e0f2f1", "100": "#b2dfdb", "200": "#80cbc4", "300": "#4db6ac", "400": "#26a69a", "500": "#009688", "600": "#00897b", "700": "#00796b", "800": "#00695c", "900": "#004d40"},
    "green": {"50": "#e8f5e9", "100": "#c8e6c9", "200": "#a5d6a7", "300": "#81c784", "400": "#66bb6a", "500": "#4caf50", "600": "#43a047", "700": "#388e3c", "800": "#2e7d32", "900": "#1b5e20"},
    "yellow": {"50": "#fffde7", "100": "#fff9c4", "200": "#fff59d", "300": "#fff176", "400": "#ffee58", "500": "#ffeb3b", "600": "#fdd835", "700": "#fbc02d", "800": "#f9a825", "900": "#f57f17"},
    "orange": {"50": "#fff3e0", "100": "#ffe0b2", "200": "#ffcc80", "300": "#ffb74d", "400": "#ffa726", "500": "#ff9800", "600": "#fb8c00", "700": "#f57c00", "800": "#ef6c00", "900": "#e65100"},
}

# Ant Design Colors - Asian Market Optimized (90K+ stars)
ANT_DESIGN_COLORS = {
    "blue": {"1": "#e6f4ff", "2": "#bae0ff", "3": "#91caff", "4": "#69b1ff", "5": "#4096ff", "6": "#1677ff", "7": "#0958d9", "8": "#003eb3", "9": "#002c8c", "10": "#001d66"},
    "purple": {"1": "#f9f0ff", "2": "#efdbff", "3": "#d3adf7", "4": "#b37feb", "5": "#9254de", "6": "#722ed1", "7": "#531dab", "8": "#391085", "9": "#22075e", "10": "#120338"},
    "cyan": {"1": "#e6fffb", "2": "#b5f5ec", "3": "#87e8de", "4": "#5cdbd3", "5": "#22c1c3", "6": "#00a4a4", "7": "#00898b", "8": "#006b75", "9": "#00505b", "10": "#003442"},
    "green": {"1": "#f6ffed", "2": "#d9f7be", "3": "#b7eb8f", "4": "#95de64", "5": "#73d13d", "6": "#52c41a", "7": "#389e0d", "8": "#237804", "9": "#135200", "10": "#092b00"},
    "red": {"1": "#fff1f0", "2": "#ffccc7", "3": "#ffa39e", "4": "#ff7875", "5": "#ff4d4f", "6": "#f5222d", "7": "#cf1322", "8": "#a8071a", "9": "#7a0413", "10": "#5c0011"},
    "orange": {"1": "#fff7e6", "2": "#ffe7ba", "3": "#ffd591", "4": "#ffc069", "5": "#ffa940", "6": "#fa8c16", "7": "#d46b08", "8": "#ad4e00", "9": "#873800", "10": "#612500"},
    "yellow": {"1": "#feffe6", "2": "#ffffb8", "3": "#fffb8f", "4": "#fff566", "5": "#ffec3d", "6": "#fadb14", "7": "#d4b106", "8": "#ad8b00", "9": "#876800", "10": "#614700"},
}

# IBM Carbon Colors - Enterprise Professional (8K+ stars)
IBM_CARBON_COLORS = {
    "blue": {"10": "#edf5ff", "20": "#d0e2ff", "30": "#a6c8ff", "40": "#78a9ff", "50": "#4589ff", "60": "#0f62fe", "70": "#0043ce", "80": "#002d9c", "90": "#001d6c", "100": "#001141"},
    "cyan": {"10": "#e5f9ff", "20": "#baefff", "30": "#82ddff", "40": "#33cbff", "50": "#00b3ff", "60": "#0092ff", "70": "#0072c3", "80": "#005387", "90": "#003552", "100": "#001a29"},
    "gray": {"10": "#f6f6f6", "20": "#e5e5e5", "30": "#c6c6c6", "40": "#a8a8a8", "50": "#8d8d8d", "60": "#6f6f6f", "70": "#525252", "80": "#393939", "90": "#262626", "100": "#161616"},
    "green": {"10": "#defbe6", "20": "#a7f4c4", "30": "#6ee3a0", "40": "#37d07b", "50": "#19c060", "60": "#0ba947", "70": "#078a38", "80": "#056b2b", "90": "#034d1e", "100": "#013013"},
    "red": {"10": "#fff1f1", "20": "#ffd7d9", "30": "#ffb3b8", "40": "#ff7d7f", "50": "#ff4d4f", "60": "#da1e28", "70": "#a2191f", "80": "#750e13", "90": "#520408", "100": "#2d0304"},
}

# GitHub Primer Colors - Data Visualization Optimized (3K+ stars)
GITHUB_PRIMER_COLORS = {
    "coral": {"1": "#fff5f1", "2": "#ffe6d5", "3": "#ffd0b5", "4": "#ffb48c", "5": "#ff925f", "6": "#ff6f2e", "7": "#e85a1a", "8": "#bc4413", "9": "#94320e", "10": "#75250a"},
    "gray": {"1": "#f6f8fa", "2": "#eaeef2", "3": "#d0d7de", "4": "#afb8c1", "5": "#8c959f", "6": "#6e7781", "7": "#57606a", "8": "#424a53", "9": "#32383f", "10": "#25292e"},
    "blue": {"1": "#ddf4ff", "2": "#b6e3ff", "3": "#80ccff", "4": "#54aeff", "5": "#218bff", "6": "#0969da", "7": "#0550ae", "8": "#033d8b", "9": "#0a3069", "10": "#001a47"},
    "green": {"1": "#dafbe1", "2": "#aceebb", "3": "#6fdd8b", "4": "#4ac26b", "5": "#2da44e", "6": "#1a7f37", "7": "#116329", "8": "#044f1e", "9": "#003d16", "10": "#002d11"},
    "orange": {"1": "#fff1e5", "2": "#ffd8b5", "3": "#ffb775", "4": "#ff8d32", "5": "#fd7504", "6": "#df5c02", "7": "#b54702", "8": "#8d3504", "9": "#692107", "10": "#4f170b"},
    "pink": {"1": "#ffeef8", "2": "#fad1df", "3": "#f7a8c7", "4": "#f575ad", "5": "#ef4e8a", "6": "#d0356e", "7": "#a82858", "8": "#821c45", "9": "#5e1433", "10": "#420e25"},
    "purple": {"1": "#f7f1ff", "2": "#e9d7fe", "3": "#d8b9fe", "4": "#c297ff", "5": "#a475f9", "6": "#8957e5", "7": "#6e40c9", "8": "#55309e", "9": "#3f2374", "10": "#2d1a52"},
    "red": {"1": "#ffebe9", "2": "#ffcecb", "3": "#ffaba8", "4": "#ff8182", "5": "#ff4f4a", "6": "#da3633", "7": "#b62324", "8": "#8e1519", "9": "#670a0e", "10": "#4b090b"},
}

# Open Color - Scandinavian Design (Original Request)
OPEN_COLOR = {
    "gray": ["#f8f9fa", "#f1f3f5", "#e9ecef", "#dee2e6", "#ced4da", "#adb5bd", "#868e96", "#495057", "#343a40", "#212529"],
    "red": ["#fff5f5", "#ffe3e3", "#ffc9c9", "#ffa8a8", "#ff8787", "#ff6b6b", "#fa5252", "#f03e3e", "#e03131", "#c92a2a"],
    "pink": ["#fff0f6", "#ffdeeb", "#fcc2d7", "#faa2c1", "#f783ac", "#f06595", "#e64980", "#d6336c", "#c2255c", "#a61e4d"],
    "grape": ["#f8f0fc", "#f3d9fa", "#eebefa", "#e599f7", "#da77f2", "#cc5de8", "#be4bdb", "#ae3ec9", "#9c36b5", "#862e9c"],
    "violet": ["#f3f0ff", "#e5dbff", "#d0bfff", "#b197fc", "#9775fa", "#845ef7", "#7950f2", "#7048e8", "#6741d9", "#5f3dc4"],
    "indigo": ["#edf2ff", "#dbe4ff", "#bac8ff", "#91a7ff", "#748ffc", "#5c7cfa", "#4c6ef5", "#4263eb", "#3b5bdb", "#364fc7"],
    "blue": ["#e7f5ff", "#d0ebff", "#a5d8ff", "#74c0fc", "#4dabf7", "#339af0", "#228be6", "#1c7ed6", "#1971c2", "#1864ab"],
    "cyan": ["#e3fafc", "#c5f6fa", "#99e9f2", "#66d9e8", "#3bc9db", "#22b8cf", "#15aabf", "#1098ad", "#0c8599", "#0b7285"],
    "teal": ["#e6fcf5", "#c3fae8", "#96f2d7", "#63e6be", "#38d9a9", "#20c997", "#12b886", "#0ca678", "#099268", "#087f5b"],
    "green": ["#ebfbee", "#d3f9d8", "#b2f2bb", "#8ce99a", "#69db7c", "#51cf66", "#40c057", "#37b24d", "#2f9e44", "#2b8a3e"],
    "lime": ["#f4fce3", "#e9fac8", "#d8f5a2", "#c0eb75", "#a9e34b", "#94d82d", "#82c91e", "#74b816", "#66a80f", "#5c940d"],
    "yellow": ["#fff9db", "#fff3bf", "#ffec99", "#ffe066", "#ffd43b", "#fcc419", "#fab005", "#f59f00", "#f08c00", "#e67700"],
    "orange": ["#fff4e6", "#ffe8cc", "#ffd8a8", "#ffc078", "#ffa94d", "#ff922b", "#fd7e14", "#f76707", "#e8590c", "#d9480f"],
}

# Consolidated Design Tokens
DESIGN_TOKENS = {
    "colors": {
        **BRAND_COLORS,
        "_systems": {
            "tailwind": TAILWIND_COLORS,
            "material": MATERIAL_COLORS,
            "ant_design": ANT_DESIGN_COLORS,
            "ibm_carbon": IBM_CARBON_COLORS,
            "github_primer": GITHUB_PRIMER_COLORS,
            "open_color": OPEN_COLOR,
        }
    },
    "chart_colors": [
        "#1a5f3c", "#0d3d26", "#ffb1ee", "#3b82f6", "#ef4444", 
        "#f59e0b", "#10b981", "#8b5cf6", "#ec4899", "#06b6d4"
    ]
}


def get_color(system: str, color_name: str, shade: str = "500") -> str:
    """
    Get color from any integrated design system.
    
    Args:
        system: Design system name (tailwind, material, ant_design, ibm_carbon, 
                github_primer, open_color)
        color_name: Color family name (blue, red, green, etc.)
        shade: Shade level (varies by system: 50-900, 1-10, N10-N100, etc.)
    
    Returns:
        Hex color string or None if not found
    """
    systems_map = DESIGN_TOKENS["colors"]["_systems"]
    if system not in systems_map:
        return BRAND_COLORS["primary"]
    
    palette = systems_map[system]
    if color_name not in palette:
        return BRAND_COLORS["primary"]
    
    shades = palette[color_name]
    if isinstance(shades, dict):
        return shades.get(str(shade), BRAND_COLORS["primary"])
    elif isinstance(shades, list):
        try:
            idx = int(shade) if shade.isdigit() else 5
            return shades[min(idx, len(shades)-1)]
        except:
            return shades[5] if len(shades) > 5 else shades[0]
    
    return BRAND_COLORS["primary"]


def run(payload: str, template: str = "default", color_system: str = "tailwind") -> str:
    """
    Payload format: filename:col1,col2,col3
    Example: Inventory:Item,Qty,Price
    
    Templates available:
    - default: Clean professional spreadsheet
    - report: Business report with styled headers
    - data: Data-heavy sheet with alternating rows
    - minimal: Sparse, focused data
    
    Color systems: tailwind, material, ant_design, ibm_carbon, github_primer, open_color
    
    Args:
        payload: String containing filename and column definitions
        template: Template style to apply
        color_system: Design system for colors
    
    Returns:
        Status string
    """
    try:
        parts = payload.split(":", 1)
        filename = parts[0].strip() if parts[0] else "output"
        columns_raw = parts[1].strip() if len(parts) > 1 else "Data"
        columns = [c.strip() for c in columns_raw.split(",")]

        # Create DataFrame with sample data structure
        df = pd.DataFrame(columns=columns)
        path = f"{filename}.xlsx"
        
        # Use openpyxl for advanced styling
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"[:31]  # Excel sheet name limit
        
        # Apply template styling
        if template == "report":
            _apply_report_template(ws, df, color_system)
        elif template == "data":
            _apply_data_template(ws, df, color_system)
        elif template == "minimal":
            _apply_minimal_template(ws, df, color_system)
        else:  # default
            _apply_default_template(ws, df, color_system)
        
        wb.save(path)
        return f"✅ Excel created: {path} ({template} template) | Columns: {', '.join(columns)}"

    except Exception as e:
        return f"❌ Excel failed: {str(e)}"


def _hex_to_rgb(hex_color: str) -> tuple:
    """Convert hex color to RGB tuple."""
    hex_color = hex_color.lstrip('#')
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))


def _apply_default_template(ws, df, color_system: str) -> None:
    """Clean professional spreadsheet with subtle branding."""
    # Header row styling
    header_fill = PatternFill(start_color=_hex_to_rgb(get_color(color_system, "blue", "600")),
                              end_color=_hex_to_rgb(get_color(color_system, "blue", "600")),
                              fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col_num, column_name in enumerate(df.columns.tolist(), 1):
        cell = ws.cell(row=1, column=col_num, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Column widths
    for col_num in range(1, len(df.columns) + 1):
        ws.column_dimensions[chr(64 + col_num)].width = 15


def _apply_report_template(ws, df, color_system: str) -> None:
    """Business report with styled headers and borders."""
    # Header styling
    header_fill = PatternFill(start_color=_hex_to_rgb(get_color(color_system, "blue", "700")),
                              end_color=_hex_to_rgb(get_color(color_system, "blue", "700")),
                              fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin', color=_hex_to_rgb(get_color(color_system, "gray", "300"))),
        right=Side(style='thin', color=_hex_to_rgb(get_color(color_system, "gray", "300"))),
        top=Side(style='thin', color=_hex_to_rgb(get_color(color_system, "gray", "300"))),
        bottom=Side(style='thin', color=_hex_to_rgb(get_color(color_system, "gray", "300")))
    )
    
    # Add headers
    for col_num, column_name in enumerate(df.columns.tolist(), 1):
        cell = ws.cell(row=1, column=col_num, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Column widths
    for col_num in range(1, len(df.columns) + 1):
        ws.column_dimensions[chr(64 + col_num)].width = 18


def _apply_data_template(ws, df, color_system: str) -> None:
    """Data-heavy sheet with alternating row colors."""
    # Header styling
    header_fill = PatternFill(start_color=_hex_to_rgb(get_color(color_system, "slate", "700")),
                              end_color=_hex_to_rgb(get_color(color_system, "slate", "700")),
                              fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    
    # Alternating row fills
    odd_fill = PatternFill(start_color=_hex_to_rgb(get_color(color_system, "gray", "50")),
                           end_color=_hex_to_rgb(get_color(color_system, "gray", "50")),
                           fill_type="solid")
    even_fill = PatternFill(start_color="FFFFFF",
                            end_color="FFFFFF",
                            fill_type="solid")
    
    # Add headers
    for col_num, column_name in enumerate(df.columns.tolist(), 1):
        cell = ws.cell(row=1, column=col_num, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Column widths
    for col_num in range(1, len(df.columns) + 1):
        ws.column_dimensions[chr(64 + col_num)].width = 16


def _apply_minimal_template(ws, df, color_system: str) -> None:
    """Sparse, focused data with maximum whitespace."""
    # Minimal header styling
    header_font = Font(bold=True, color=_hex_to_rgb(get_color(color_system, "gray", "700")), size=11)
    
    # Add headers
    for col_num, column_name in enumerate(df.columns.tolist(), 1):
        cell = ws.cell(row=1, column=col_num, value=column_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="bottom")
    
    # Column widths
    for col_num in range(1, len(df.columns) + 1):
        ws.column_dimensions[chr(64 + col_num)].width = 14
